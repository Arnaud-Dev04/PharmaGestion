"""
Report Service - Generates Excel and PDF reports for system data.
"""

from typing import Optional
from datetime import date, datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from sqlalchemy import func, desc

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from app.models.medicine import Medicine
from app.models.sales import Sale, SaleItem



def _create_excel_header(ws, headers):
    """Helper to create styled header row in Excel."""
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")


def generate_stock_excel(db: Session) -> BytesIO:
    """
    Generate Excel file with current stock status.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport de Stock"

    # Headers
    headers = ["ID", "Code", "Nom", "Quantité", "Alerte Min", "P. Achat", "P. Vente", "Expiration"]
    _create_excel_header(ws, headers)

    # Data
    medicines = db.query(Medicine).order_by(Medicine.name).all()
    for row_num, med in enumerate(medicines, 2):
        ws.cell(row=row_num, column=1, value=med.id)
        ws.cell(row=row_num, column=2, value=med.code)
        ws.cell(row=row_num, column=3, value=med.name)
        ws.cell(row=row_num, column=4, value=med.quantity)
        ws.cell(row=row_num, column=5, value=med.min_stock_alert)
        ws.cell(row=row_num, column=6, value=med.price_buy)
        ws.cell(row=row_num, column=7, value=med.price_sell)
        ws.cell(row=row_num, column=8, value=med.expiry_date)

    # Auto-adjust column width (simple)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_sales_excel(db: Session, start_date: date, end_date: date) -> BytesIO:
    """
    Generate Excel file with sales history filtered by date.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport des Ventes"

    # Headers
    headers = ["Facture", "Date", "Vendeur", "Montant Total", "Articles", "Paiement"]
    _create_excel_header(ws, headers)

    # Data
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())

    sales = db.query(Sale).filter(
        Sale.date >= start_dt,
        Sale.date <= end_dt
    ).order_by(Sale.date.desc()).all()

    for row_num, sale in enumerate(sales, 2):
        payment_method = sale.payment_method
        if hasattr(payment_method, 'value'):
             payment_method = payment_method.value
             
        ws.cell(row=row_num, column=1, value=sale.code)
        ws.cell(row=row_num, column=2, value=sale.date)
        ws.cell(row=row_num, column=3, value=sale.user.username if sale.user else "Inconnu")
        ws.cell(row=row_num, column=4, value=sale.total_amount)
        ws.cell(row=row_num, column=5, value=len(sale.items))
        ws.cell(row=row_num, column=6, value=payment_method)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_financial_pdf(db: Session, start_date: Optional[date] = None, end_date: Optional[date] = None, period_label: str = "Aperçu") -> BytesIO:
    """
    Generate a professional PDF financial summary with company header.
    Groups medicines by family with quantity, rate (price), and value.
    Format inspired by MUHIRWE PHARMA but branded as PHARMA-SOURCE.
    """
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    
    buffer = BytesIO()
    
    # Create canvas with A4 size
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Constants for layout
    margin = 2 * cm
    y_position = height - margin
    
    # Company Header
    c.setFont("Helvetica-Bold", 14)
    c.drawString(margin, y_position, "PHARMA-SOURCE")
    y_position -= 0.5 * cm
    
    c.setFont("Helvetica", 10)
    c.drawString(margin, y_position, "NIF: [À compléter]")  # User will provide NIF later
    y_position -= 0.4 * cm
    c.drawString(margin, y_position, "TEL: [À compléter]")  # User will provide TEL later
    y_position -= 1 * cm
    
    # Report Title and Date
    c.setFont("Helvetica-Bold", 12)
    if start_date and end_date:
        date_str = f"Du {start_date.strftime('%d-%b-%Y')} au {end_date.strftime('%d-%b-%Y')}"
    else:
        date_str = f"Au {date.today().strftime('%d-%b-%Y')}"
    
    # Right-aligned company name and date
    company_text = "PHARMA-SOURCE"
    c.drawRightString(width - margin, y_position + 0.5 * cm, company_text)
    c.setFont("Helvetica", 10)
    c.drawRightString(width - margin, y_position, f"For {date_str}")
    y_position -= 1.2 * cm
    
    # Table Headers
    c.setFont("Helvetica-Bold", 10)
    c.drawString(margin, y_position, "Particuliers")
    c.drawCentredString(width - margin - 9*cm, y_position, "Closing Balance")
    
    # Sub-headers for Closing Balance section
    c.setFont("Helvetica", 9)
    col_quantity_x = width - margin - 12*cm
    col_rate_x = width - margin - 8*cm  
    col_value_x = width - margin - 3*cm
    
    y_position -= 0.5 * cm
    c.drawString(col_quantity_x, y_position, "Quantité")
    c.drawString(col_rate_x, y_position, "Taux")
    c.drawString(col_value_x, y_position, "Valeur")
    
    # Draw header line
    y_position -= 0.2 * cm
    c.line(margin, y_position, width - margin, y_position)
    y_position -= 0.5 * cm
    
    # Fetch and group medicines by family
    from app.models.medicine import MedicineFamily
    
    # Build query based on date filter if provided
    if start_date and end_date:
        # Get medicines that were sold in the period
        start_dt = datetime.combine(start_date, datetime.min.time())
        end_dt = datetime.combine(end_date, datetime.max.time())
        
        # Query to get sold medicines with quantities
        sold_medicines = db.query(
            Medicine.id,
            Medicine.name,
            Medicine.family_id,
            Medicine.price_sell,
            func.sum(SaleItem.quantity).label('quantity_sold')
        ).join(SaleItem).join(Sale).filter(
            Sale.date >= start_dt,
            Sale.date <= end_dt
        ).group_by(Medicine.id).all()
        
        # Organize by family
        medicines_by_family = {}
        for med_id, med_name, family_id, price_sell, qty_sold in sold_medicines:
            if family_id not in medicines_by_family:
                family = db.query(MedicineFamily).filter(MedicineFamily.id == family_id).first()
                family_name = family.name if family else "Sans Catégorie"
                medicines_by_family[family_id] = {
                    'name': family_name,
                    'items': []
                }
            medicines_by_family[family_id]['items'].append({
                'name': med_name,
                'quantity': qty_sold,
                'rate': price_sell,
                'value': qty_sold * price_sell
            })
    else:
        # Get current stock
        medicines = db.query(Medicine).filter(Medicine.is_active == True).all()
        
        medicines_by_family = {}
        for med in medicines:
            family_id = med.family_id or 0
            if family_id not in medicines_by_family:
                family_name = med.family.name if med.family else "Sans Catégorie"
                medicines_by_family[family_id] = {
                    'name': family_name,
                    'items': []
                }
            medicines_by_family[family_id]['items'].append({
                'name': med.name,
                'quantity': med.quantity,
                'rate': med.price_sell,
                'value': med.quantity * med.price_sell
            })
    
    # Draw data rows
    c.setFont("Helvetica", 9)
    grand_total_value = 0
    
    for family_data in medicines_by_family.values():
        # Check if we need a new page
        if y_position < 3 * cm:
            c.showPage()
            y_position = height - margin
            c.setFont("Helvetica", 9)
        
        # Family header
        c.setFont("Helvetica-Bold", 9)
        c.drawString(margin, y_position, family_data['name'])
        y_position -= 0.4 * cm
        
        # Family items
        c.setFont("Helvetica", 8)
        for item in family_data['items']:
            if y_position < 3 * cm:
                c.showPage()
                y_position = height - margin
                c.setFont("Helvetica", 8)
            
            # Item name (truncated if too long)
            item_name = item['name'][:40]
            c.drawString(margin + 0.5*cm, y_position, item_name)
            
            # Quantity, Rate, Value
            c.drawRightString(col_quantity_x + 2*cm, y_position, f"{item['quantity']:.0f}")
            c.drawRightString(col_rate_x + 2*cm, y_position, f"{item['rate']:.2f}")
            c.drawRightString(col_value_x + 3*cm, y_position, f"{item['value']:.2f}")
            
            grand_total_value += item['value']
            y_position -= 0.35 * cm
        
        y_position -= 0.3 * cm  # Space after family
    
    # Grand Total
    y_position -= 0.5 * cm
    c.line(margin, y_position, width - margin, y_position)
    y_position -= 0.5 * cm
    
    c.setFont("Helvetica-Bold", 10)
    c.drawString(margin, y_position, "Grand Total")
    c.drawRightString(col_value_x + 3*cm, y_position, f"{grand_total_value:.2f}")
    
    # Footer - space for signature and stamp
    y_position = 3 * cm
    c.setFont("Helvetica-Italic", 8)
    c.drawCentredString(width / 2, y_position, "Signature & Cachet")
    
    c.save()
    buffer.seek(0)
    return buffer



def generate_stock_pdf(db: Session) -> BytesIO:
    """
    Generate professional PDF file with current stock status.
    Format: PHARMA-SOURCE branded with company header.
    """
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas
    
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    margin = 2 * cm
    y_position = height - margin
    
    # Company Header
    c.setFont("Helvetica-Bold", 14)
    c.drawString(margin, y_position, "PHARMA-SOURCE")
    y_position -= 0.5 * cm
    
    c.setFont("Helvetica", 10)
    c.drawString(margin, y_position, "NIF: [À compléter]")
    y_position -= 0.4 * cm
    c.drawString(margin, y_position, "TEL: [À compléter]")
    y_position -= 1 * cm
    
    # Report Title and Date - Right aligned
    c.setFont("Helvetica-Bold", 12)
    c.drawRightString(width - margin, y_position + 0.5 * cm, "PHARMA-SOURCE")
    c.setFont("Helvetica", 10)
    c.drawRightString(width - margin, y_position, f"État du Stock - {date.today().strftime('%d-%b-%Y')}")
    y_position -= 1.5 * cm
    
    # Table Headers
    c.setFont("Helvetica-Bold", 9)
    col_code_x = margin
    col_name_x = margin + 3*cm
    col_qty_x = width - margin - 9*cm
    col_min_x = width - margin - 7*cm
    col_buy_x = width - margin - 5*cm
    col_sell_x = width - margin - 3*cm
    col_exp_x = width - margin - 1*cm - 1.5*cm
    
    c.drawString(col_code_x, y_position, "Code")
    c.drawString(col_name_x, y_position, "Nom")
    c.drawString(col_qty_x, y_position, "Qté")
    c.drawString(col_min_x, y_position, "Min")
    c.drawString(col_buy_x, y_position, "P.A")
    c.drawString(col_sell_x, y_position, "P.V")
    c.drawString(col_exp_x, y_position, "Exp.")
    
    # Header line
    y_position -= 0.2 * cm
    c.line(margin, y_position, width - margin, y_position)
    y_position -= 0.5 * cm
    
    # Fetch data
    medicines = db.query(Medicine).filter(Medicine.is_active == True).order_by(Medicine.name).all()
    
    c.setFont("Helvetica", 8)
    row_count = 0
    
    for med in medicines:
        if y_position < 3 * cm:
            c.showPage()
            y_position = height - margin
            c.setFont("Helvetica", 8)
        
        # Highlight low stock
        if med.quantity <= med.min_stock_alert:
            c.setFillColorRGB(1, 0.8, 0.8)  # Light red background
            c.rect(margin - 0.2*cm, y_position - 0.1*cm, width - 2*margin + 0.4*cm, 0.4*cm, fill=1)
            c.setFillColorRGB(0, 0, 0)  # Reset to black text
        
        c.drawString(col_code_x, y_position, med.code[:10])
        c.drawString(col_name_x, y_position, med.name[:25])
        c.drawRightString(col_qty_x + 1*cm, y_position, f"{med.quantity:.0f}")
        c.drawRightString(col_min_x + 1*cm, y_position, f"{med.min_stock_alert}")
        c.drawRightString(col_buy_x + 1*cm, y_position, f"{med.price_buy:.0f}")
        c.drawRightString(col_sell_x + 1*cm, y_position, f"{med.price_sell:.0f}")
        
        exp_str = med.expiry_date.strftime("%d/%m/%y") if med.expiry_date else "-"
        c.drawString(col_exp_x, y_position, exp_str)
        
        y_position -= 0.35 * cm
        row_count += 1
    
    # Summary
    y_position -= 0.5 * cm
    c.line(margin, y_position, width - margin, y_position)
    y_position -= 0.5 * cm
    
    c.setFont("Helvetica-Bold", 9)
    c.drawString(margin, y_position, f"Total Médicaments: {len(medicines)}")
    
    # Footer
    y_position = 2.5 * cm
    c.setFont("Helvetica-Italic", 8)
    c.drawCentredString(width / 2, y_position, "Signature & Cachet")
    
    c.save()
    buffer.seek(0)
    return buffer


def generate_sales_pdf(db: Session, start_date: date, end_date: date) -> BytesIO:
    """
    Generate professional PDF file with sales history.
    Format: PHARMA-SOURCE branded with company header.
    """
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas
    
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    margin = 2 * cm
    y_position = height - margin
    
    # Company Header
    c.setFont("Helvetica-Bold", 14)
    c.drawString(margin, y_position, "PHARMA-SOURCE")
    y_position -= 0.5 * cm
    
    c.setFont("Helvetica", 10)
    c.drawString(margin, y_position, "NIF: [À compléter]")
    y_position -= 0.4 * cm
    c.drawString(margin, y_position, "TEL: [À compléter]")
    y_position -= 1 * cm
    
    # Report Title and Date - Right aligned
    c.setFont("Helvetica-Bold", 12)
    c.drawRightString(width - margin, y_position + 0.5 * cm, "PHARMA-SOURCE")
    c.setFont("Helvetica", 10)
    date_range = f"Du {start_date.strftime('%d-%b-%Y')} au {end_date.strftime('%d-%b-%Y')}"
    c.drawRightString(width - margin, y_position, f"Rapport des Ventes - {date_range}")
    y_position -= 1.5 * cm
    
    # Table Headers
    c.setFont("Helvetica-Bold", 9)
    col_invoice_x = margin
    col_date_x = margin + 3*cm
    col_user_x = margin + 6*cm
    col_amount_x = width - margin - 5*cm
    col_items_x = width - margin - 3*cm
    col_payment_x = width - margin - 1.5*cm
    
    c.drawString(col_invoice_x, y_position, "Facture")
    c.drawString(col_date_x, y_position, "Date")
    c.drawString(col_user_x, y_position, "Vendeur")
    c.drawString(col_amount_x, y_position, "Montant")
    c.drawString(col_items_x, y_position, "Arts")
    c.drawString(col_payment_x, y_position, "Pmt")
    
    # Header line
    y_position -= 0.2 * cm
    c.line(margin, y_position, width - margin, y_position)
    y_position -= 0.5 * cm
    
    # Fetch data
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())
    
    sales = db.query(Sale).filter(
        Sale.date >= start_dt,
        Sale.date <= end_dt
    ).order_by(Sale.date.desc()).all()
    
    c.setFont("Helvetica", 8)
    total_period = 0.0
    
    for sale in sales:
        if y_position < 4 * cm:
            c.showPage()
            y_position = height - margin
            c.setFont("Helvetica", 8)
        
        payment_method = sale.payment_method
        if hasattr(payment_method, 'value'):
            payment_method = payment_method.value
        
        c.drawString(col_invoice_x, y_position, sale.code[:12])
        c.drawString(col_date_x, y_position, sale.date.strftime("%d/%m/%y"))
        username = sale.user.username[:12] if sale.user else "N/A"
        c.drawString(col_user_x, y_position, username)
        c.drawRightString(col_amount_x + 1.5*cm, y_position, f"{sale.total_amount:,.0f}")
        c.drawString(col_items_x, y_position, str(len(sale.items)))
        pmt_short = str(payment_method)[:8] if payment_method else "-"
        c.drawString(col_payment_x, y_position, pmt_short)
        
        total_period += sale.total_amount
        y_position -= 0.35 * cm
    
    # Grand Total
    y_position -= 0.5 * cm
    c.line(margin, y_position, width - margin, y_position)
    y_position -= 0.5 * cm
    
    c.setFont("Helvetica-Bold", 10)
    c.drawString(margin, y_position, "TOTAL")
    c.drawRightString(col_amount_x + 1.5*cm, y_position, f"{total_period:,.2f} FBu")
    
    y_position -= 0.5 * cm
    c.setFont("Helvetica", 9) 
    c.drawString(margin, y_position, f"Nombre de ventes: {len(sales)}")
    
    # Footer
    y_position = 2.5 * cm
    c.setFont("Helvetica-Italic", 8)
    c.drawCentredString(width / 2, y_position, "Signature & Cachet")
    
    c.save()
    buffer.seek(0)
    return buffer


def generate_stock_word(db: Session) -> BytesIO:
    """
    Generate Word file (MHTML/HTML compatible) with current stock status.
    """
    medicines = db.query(Medicine).order_by(Medicine.name).all()
    
    html = f"""
    <html xmlns:o='urn:schemas-microsoft-com:office:office' xmlns:w='urn:schemas-microsoft-com:office:word' xmlns='http://www.w3.org/TR/REC-html40'>
    <head>
        <meta charset="utf-8">
        <title>État du Stock</title>
        <style>
            body {{ font-family: Arial, sans-serif; }}
            table {{ border-collapse: collapse; width: 100%; }}
            th, td {{ border: 1px solid black; padding: 8px; text-align: left; }}
            th {{ background-color: #4F81BD; color: white; }}
            .header {{ font-size: 18px; font-weight: bold; margin-bottom: 20px; }}
            .date {{ margin-bottom: 20px; }}
        </style>
    </head>
    <body>
        <div class="header">État du Stock</div>
        <div class="date">Généré le: {date.today().strftime('%d/%m/%Y')}</div>
        <table>
            <thead>
                <tr>
                    <th>Code</th>
                    <th>Nom</th>
                    <th>Quantité</th>
                    <th>P. Vente</th>
                    <th>Expiration</th>
                </tr>
            </thead>
            <tbody>
    """
    
    for med in medicines:
        html += f"""
                <tr>
                    <td>{med.code}</td>
                    <td>{med.name}</td>
                    <td>{med.quantity}</td>
                    <td>{med.price_sell:,.0f}</td>
                    <td>{med.expiry_date.strftime("%d/%m/%Y") if med.expiry_date else "-"}</td>
                </tr>
        """
        
    html += """
            </tbody>
        </table>
    </body>
    </html>
    """
    
    return BytesIO(html.encode('utf-8'))


def generate_sales_word(db: Session, start_date: date, end_date: date) -> BytesIO:
    """
    Generate Word file (MHTML/HTML compatible) with sales history.
    """
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())

    sales = db.query(Sale).filter(
        Sale.date >= start_dt,
        Sale.date <= end_dt
    ).order_by(Sale.date.desc()).all()
    
    total_period = sum(s.total_amount for s in sales)

    html = f"""
    <html xmlns:o='urn:schemas-microsoft-com:office:office' xmlns:w='urn:schemas-microsoft-com:office:word' xmlns='http://www.w3.org/TR/REC-html40'>
    <head>
        <meta charset="utf-8">
        <title>Rapport des Ventes</title>
        <style>
            body {{ font-family: Arial, sans-serif; }}
            table {{ border-collapse: collapse; width: 100%; }}
            th, td {{ border: 1px solid black; padding: 8px; text-align: center; }}
            th {{ background-color: #C0504D; color: white; }}
            .header {{ font-size: 18px; font-weight: bold; margin-bottom: 20px; }}
            .period {{ margin-bottom: 20px; }}
            .total-row {{ font-weight: bold; background-color: #D3D3D3; }}
        </style>
    </head>
    <body>
        <div class="header">Rapport des Ventes</div>
        <div class="period">Période: {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}</div>
        <table>
            <thead>
                <tr>
                    <th>Facture</th>
                    <th>Date</th>
                    <th>Total (FBu)</th>
                    <th>Articles</th>
                    <th>Paiement</th>
                </tr>
            </thead>
            <tbody>
    """
    
    for sale in sales:
        payment_method = sale.payment_method
        if hasattr(payment_method, 'value'):
             payment_method = payment_method.value
             
        html += f"""
                <tr>
                    <td>{sale.code}</td>
                    <td>{sale.date.strftime("%d/%m/%Y")}</td>
                    <td>{sale.total_amount:,.0f}</td>
                    <td>{len(sale.items)}</td>
                    <td>{payment_method}</td>
                </tr>
        """
        
    html += f"""
                <tr class="total-row">
                    <td colspan="2">TOTAL</td>
                    <td>{total_period:,.0f} FBu</td>
                    <td colspan="2"></td>
                </tr>
            </tbody>
        </table>
    </body>
    </html>
    """
    
    return BytesIO(html.encode('utf-8'))
